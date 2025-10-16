import streamlit as st
import pandas as pd
import re
import io
from openpyxl import load_workbook
from openpyxl.comments import Comment

# ---------------- Helpers ----------------
def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def clean_room_code(building, floor, room):
    """Smart room code extraction logic based on patterns."""
    if not isinstance(room, str) or not room.strip():
        return ""

    room = room.strip()
    building_suffix = str(building).split("-")[-1] if isinstance(building, str) else ""
    floor_str = str(floor).strip()

    # Case 1: Hyphen-separated pattern
    if "-" in room:
        parts = [p.strip() for p in room.split("-") if p.strip()]
        if len(parts) >= 3:
            if (
                parts[0].endswith(building_suffix)
                or parts[1] == floor_str
                or floor_str in parts
            ):
                return parts[-1]
        return parts[-1] if parts else room

    # Case 2: Dot-separated pattern
    if "." in room:
        parts = [p.strip() for p in room.split(".") if p.strip()]
        return parts[-1] if parts else room

    # Case 3: No recognizable pattern
    return room


def extract_term_abbrev_table(tag_summary_raw: pd.DataFrame) -> pd.DataFrame:
    """Extract Term ‚Üî Abbreviation pairs from Tag Summary."""
    df = tag_summary_raw.copy().astype(str).applymap(lambda x: x.strip() if pd.notna(x) else "")
    rows, cols = df.shape
    term_pos = abbr_pos = None

    for r in range(rows):
        for c in range(cols):
            cell = _norm(df.iat[r, c])
            if cell == "term" and term_pos is None:
                term_pos = (r, c)
            if "abbreviation" in cell and abbr_pos is None:
                abbr_pos = (r, c)

    if not term_pos or not abbr_pos:
        return pd.DataFrame(columns=["Term", "Abbreviation"])

    row_start = max(term_pos[0], abbr_pos[0]) + 1
    term_col, abbr_col = term_pos[1], abbr_pos[1]

    out_rows, blanks = [], 0
    for r in range(row_start, rows):
        term_val = df.iat[r, term_col]
        abbr_val = df.iat[r, abbr_col]

        if not term_val and not abbr_val:
            blanks += 1
        else:
            blanks = 0
        if blanks >= 3:
            break

        if _norm(term_val) in ["term", ""] or "abbreviation" in _norm(abbr_val):
            continue

        if term_val or abbr_val:
            out_rows.append({"Term": term_val, "Abbreviation": abbr_val})

    return pd.DataFrame(out_rows).drop_duplicates().reset_index(drop=True)


def extract_name_abbrev_table(sheet_raw: pd.DataFrame) -> pd.DataFrame:
    """Extract Name ‚Üî Abbreviation pairs from equipment sheets."""
    df = sheet_raw.copy().astype(str).applymap(lambda x: x.strip() if pd.notna(x) else "")
    rows, cols = df.shape
    name_pos = abbr_pos = None

    for r in range(rows):
        for c in range(cols):
            cell = _norm(df.iat[r, c])
            if ("name" in cell or "tag" in cell or "signal" in cell) and name_pos is None:
                name_pos = (r, c)
            if ("abbr" in cell or "abbreviation" in cell) and abbr_pos is None:
                abbr_pos = (r, c)

    if not name_pos or not abbr_pos:
        return pd.DataFrame(columns=["Name", "Abbreviation"])

    row_start = max(name_pos[0], abbr_pos[0]) + 1
    name_col, abbr_col = name_pos[1], abbr_pos[1]

    out_rows, blanks = [], 0
    for r in range(row_start, rows):
        name_val = df.iat[r, name_col]
        abbr_val = df.iat[r, abbr_col]

        if not name_val and not abbr_val:
            blanks += 1
        else:
            blanks = 0
        if blanks >= 5:
            break

        if "name" in _norm(name_val) or "abbreviation" in _norm(abbr_val):
            continue

        if name_val or abbr_val:
            out_rows.append({"Name": name_val, "Abbreviation": abbr_val})

    return pd.DataFrame(out_rows).drop_duplicates().reset_index(drop=True)


# ---------------- Streamlit App ----------------
st.set_page_config(page_title="Multi-Asset Nomenclature Generator", layout="wide")
st.title("üè¢ Multi-Asset Nomenclature Generator")

planon_file = st.file_uploader("üìÇ Upload DataProduct Excel", type=["xlsx"])
sys_file = st.file_uploader("üìÇ Upload Nomenclature sheet", type=["xlsx"])

if planon_file and sys_file:
    df_planon = pd.read_excel(planon_file, dtype=str)
    df_planon.columns = df_planon.columns.str.strip()

    required = ["Location Code", "Building Code", "Floor", "Rooms"]
    if not all(col in df_planon.columns for col in required):
        st.error(f"Planon must contain: {', '.join(required)}")
        st.stop()

    # Dependent dropdowns
    loc_options = sorted(df_planon["Location Code"].dropna().unique())
    location_code = st.selectbox("üåç Select Location Code", loc_options)

    building_options = sorted(df_planon.loc[df_planon["Location Code"] == location_code, "Building Code"].dropna().unique())
    building = st.selectbox("üè¢ Select Building Code", building_options)

    floor_options = sorted(
        df_planon[
            (df_planon["Location Code"] == location_code) &
            (df_planon["Building Code"] == building)
        ]["Floor"].dropna().unique()
    )
    floor = st.selectbox("üèóÔ∏è Select Floor Code", floor_options)

    room_options = sorted(
        df_planon[
            (df_planon["Location Code"] == location_code) &
            (df_planon["Building Code"] == building) &
            (df_planon["Floor"] == floor)
        ]["Rooms"].dropna().unique()
    )
    rooms = st.multiselect("üö™ Select Room Codes", room_options)

    if not rooms:
        st.warning("‚ö†Ô∏è Please select at least one room.")
        st.stop()

    # Load system workbook
    book = pd.ExcelFile(sys_file)

    # Detect Tag Summary
    tag_summary_sheet = next((s for s in book.sheet_names if re.search(r"tag\s*summary", s, re.I)), None)
    if not tag_summary_sheet:
        st.error("‚ùå No Tag Summary sheet found")
        st.stop()

    tag_raw = pd.read_excel(book, sheet_name=tag_summary_sheet, header=None, dtype=str)
    tag_df = extract_term_abbrev_table(tag_raw)

    if tag_df.empty:
        st.error("‚ùå Could not extract Term-Abbreviation pairs from Tag Summary")
        st.stop()

    equip_terms = st.multiselect("‚öôÔ∏è Select Equipments (from Tag Summary)", sorted(tag_df["Term"].dropna().unique()))

    asset_numbers = {}
    for equip in equip_terms:
        asset_numbers[equip] = st.text_input(f"üî¢ Enter Asset Number for {equip}", value="")

    if st.button("üöÄ Generate Final Nomenclatures"):
        if not equip_terms:
            st.error("‚ùå Please select at least one equipment")
            st.stop()
        if any(not num for num in asset_numbers.values()):
            st.error("‚ùå Please enter asset numbers for all selected equipments")
            st.stop()

        all_nomenclatures = []

        # üîß Correct loop structure: extract once per equipment, apply to all rooms
        for equip_term in equip_terms:
            try:
                equip_abbrev = tag_df.loc[tag_df["Term"] == equip_term, "Abbreviation"].iloc[0]
            except:
                equip_abbrev = "N/A"

            asset_number = asset_numbers.get(equip_term, "")

            if equip_term not in book.sheet_names:
                st.warning(f"‚ö†Ô∏è No sheet found for equipment '{equip_term}', skipping...")
                continue

            equip_raw = pd.read_excel(book, sheet_name=equip_term, header=None, dtype=str)
            equip_df = extract_name_abbrev_table(equip_raw)

            if equip_df.empty:
                st.warning(f"‚ö†Ô∏è Could not extract Name/Abbreviation from sheet '{equip_term}', skipping...")
                continue

            # Loop through all rooms and tags properly
            for room in rooms:
                room_clean = clean_room_code(building, floor, room)

                loc_trimmed = str(location_code).replace("LOC-", "", 1) if location_code else ""
                loc_parts = loc_trimmed.split("-")
                if len(loc_parts) >= 2:
                    loc_prefix = loc_parts[0]
                    loc_site = loc_parts[1]
                    loc_formatted = f"{loc_prefix}_{loc_site}"
                else:
                    loc_formatted = loc_trimmed

                bldg_parts = str(building).split("-", 1)
                bldg_trimmed = bldg_parts[1] if len(bldg_parts) > 1 else building
                prefix = f"{loc_formatted}_{bldg_trimmed}"
                equip_token = f"{equip_abbrev}{asset_number}"

                for _, row in equip_df.iterrows():
                    tag_name = str(row["Name"])
                    tag_abbr = str(row["Abbreviation"])
                    tag_abbr_clean = re.sub(r"\d+", "", str(tag_abbr))

                    final = f"{prefix}_{floor}_{equip_token}_{room_clean}_{tag_abbr_clean}"

                    all_nomenclatures.append([
                        location_code, building, floor, room,
                        equip_term, equip_abbrev, tag_name, tag_abbr_clean, final
                    ])

        if all_nomenclatures:
            out_df = pd.DataFrame(
                all_nomenclatures,
                columns=[
                    "Location Code", "Building code", "Floor code", "Room code",
                    "Equipment Term", "Equipment Abbreviation",
                    "Name", "Tag Abbreviation", "Final Nomenclature"
                ]
            )

            note_text = "‚ö†Ô∏è Note: Do not display number if the numbers are present in Tag Abbrevation. Numbers to be shown only for asset number."
            note_df = pd.DataFrame([[note_text] + [""] * (len(out_df.columns) - 1)], columns=out_df.columns)
            out_df = pd.concat([note_df, out_df], ignore_index=True)

            st.info(note_text)
            st.success("‚úÖ Final Nomenclatures Generated")
            st.dataframe(out_df)

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                out_df.to_excel(writer, index=False)

            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active
            note = (
                "‚ö†Ô∏è Room codes are processed using smart logic:\n"
                "‚Ä¢ Hyphen patterns ‚Üí last part after matching building/floor\n"
                "‚Ä¢ Dot patterns ‚Üí last part after final dot\n"
                "‚Ä¢ Others ‚Üí room kept as-is"
            )
            ws["I1"].comment = Comment(note, "System")

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                "üì• Download Final Excel",
                data=output,
                file_name="Final_Nomenclature_Output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("‚ùå No nomenclatures generated. Please check selections and sheet contents.")
